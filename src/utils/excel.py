import os
from openpyxl import Workbook
from src.models.product import Product

class Excel:
    def __init__(self):
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = "Products"
        self.current_row = 1
        
        # 엑셀 저장 경로 설정
        self.excel_dir = os.path.join(os.path.dirname(os.path.dirname(__file__)), "excels")
        self.excel_path = os.path.join(self.excel_dir, "extract.xlsx")
        
        # excels 디렉토리가 없으면 생성
        if not os.path.exists(self.excel_dir):
            os.makedirs(self.excel_dir)
            
        # 헤더 추가
        headers = ["제품명", "URL", "옵션", "즉시구매가", "보관판매가", "마진", "마진률", 
                  "최근거래가", "최근가격변동", "발매가", "모델번호", "출시일", "대표색상"]
        self.ws.append(headers)

    def save(self):
        try:
            self.wb.save(self.excel_path)
            print(f"엑셀 파일이 저장되었습니다: {self.excel_path}")
        except Exception as e:
            print(f"엑셀 파일 저장 중 오류 발생: {e}")

    def add_product(self, product: Product):
        """제품 정보를 엑셀에 추가합니다."""
        row_data = [
            product.name,
            product.url,
            product.option,
            product.ipp,
            product.ssp,
            product.m,
            product.mr,
            product.info.rtp,
            product.info.rpc,
            product.info.rp,
            product.info.mn,
            product.info.rd,
            product.info.rc
        ]
        
        self.ws.append(row_data)
        self.current_row += 1
        
        # 열 너비 자동 조정
        for col in range(1, len(row_data) + 1):
            column_letter = self.ws.cell(row=1, column=col).column_letter
            max_length = 0
            for row in range(1, self.current_row + 1):
                cell = self.ws.cell(row=row, column=col)
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            self.ws.column_dimensions[column_letter].width = adjusted_width


## 입력데이터의 특정입력값을 기준으로 데이터를 추출해야합니다
